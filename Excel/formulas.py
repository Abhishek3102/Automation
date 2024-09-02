from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('barchart.xlsx')
sheet = workbook['Report']

min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

for i in range(min_column+1, max_column+1):  
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

workbook.save('report.xlsx')